"""
Módulo para generar reportes en Excel de las mediciones de calidad del aire.
Los reportes incluyen TODOS los datos minuto a minuto del período seleccionado.
"""
from datetime import datetime, date, time, timedelta
from zoneinfo import ZoneInfo
from io import BytesIO
from typing import List, Optional
import logging

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

from db import db
from models import Measurement, SensorChannel
from labels import label_for

BOGOTA = ZoneInfo("America/Bogota")
logger = logging.getLogger(__name__)


def _bounds_of_day_local(d: date):
    """Retorna los límites de un día en zona local."""
    start_local = datetime.combine(d, time(0, 0, 0)).replace(tzinfo=BOGOTA)
    end_local = datetime.combine(d, time(23, 59, 59)).replace(tzinfo=BOGOTA)
    return start_local, end_local


def _bounds_of_range_local(dstart: date, dend: date):
    """Retorna los límites de un rango de fechas en zona local."""
    a, _ = _bounds_of_day_local(dstart)
    _, b = _bounds_of_day_local(dend)
    return a, b


def _get_measurements_df(
    start_dt: datetime,
    end_dt: datetime,
    devices: Optional[List[str]] = None,
    channels: Optional[List[SensorChannel]] = None
) -> pd.DataFrame:
    """
    Obtiene mediciones y las retorna como DataFrame.
    Incluye TODOS los registros del período (no solo muestra).
    """
    if channels is None:
        channels = [SensorChannel.Um1, SensorChannel.Um2]
    
    qry = Measurement.query.filter(
        Measurement.fechah_local >= start_dt,
        Measurement.fechah_local <= end_dt,
        Measurement.sensor_channel.in_(channels),
    )
    
    if devices:
        qry = qry.filter(Measurement.device_id.in_(devices))
    
    measurements = qry.order_by(Measurement.fechah_local.asc()).all()
    
    if not measurements:
        return pd.DataFrame()
    
    # Convertir a lista de diccionarios
    data = []
    for m in measurements:
        data.append({
            'Fecha/Hora': m.fechah_local.strftime('%Y-%m-%d %H:%M:%S'),
            'Dispositivo': label_for(m.device_id),
            'Device_ID': m.device_id,
            'Canal': m.sensor_channel.name,
            'PM2.5 (µg/m³)': round(m.pm25, 2) if m.pm25 is not None else None,
            'PM10 (µg/m³)': round(m.pm10, 2) if m.pm10 is not None else None,
            'Temperatura (°C)': round(m.temp, 2) if m.temp is not None else None,
            'Humedad (%)': round(m.rh, 2) if m.rh is not None else None,
            'DOY': m.doy,
            'W': round(m.w, 3) if m.w is not None else None,
        })
    
    df = pd.DataFrame(data)
    return df


def _aggregate_by_minute(df: pd.DataFrame) -> pd.DataFrame:
    """
    Agrega los datos por minuto, calculando promedios por dispositivo y canal.
    """
    if df.empty:
        return df
    
    # Convertir Fecha/Hora a datetime
    df['timestamp'] = pd.to_datetime(df['Fecha/Hora'])
    
    # Redondear a minuto
    df['minuto'] = df['timestamp'].dt.floor('min')
    
    # Agrupar por minuto, dispositivo y canal
    agg_dict = {
        'PM2.5 (µg/m³)': 'mean',
        'PM10 (µg/m³)': 'mean',
        'Temperatura (°C)': 'mean',
        'Humedad (%)': 'mean',
    }
    
    df_agg = df.groupby(['minuto', 'Dispositivo', 'Device_ID', 'Canal']).agg(agg_dict).reset_index()
    
    # Renombrar y formatear
    df_agg['Fecha/Hora'] = df_agg['minuto'].dt.strftime('%Y-%m-%d %H:%M:00')
    df_agg = df_agg.drop('minuto', axis=1)
    
    # Redondear valores
    for col in ['PM2.5 (µg/m³)', 'PM10 (µg/m³)', 'Temperatura (°C)', 'Humedad (%)']:
        df_agg[col] = df_agg[col].round(2)
    
    # Reordenar columnas
    cols = ['Fecha/Hora', 'Dispositivo', 'Device_ID', 'Canal', 'PM2.5 (µg/m³)', 
            'PM10 (µg/m³)', 'Temperatura (°C)', 'Humedad (%)']
    df_agg = df_agg[cols]
    
    return df_agg.sort_values('Fecha/Hora')


def _calculate_statistics_df(df: pd.DataFrame) -> pd.DataFrame:
    """Calcula estadísticas descriptivas del DataFrame."""
    if df.empty:
        return pd.DataFrame()
    
    stats_data = []
    variables = [
        ('PM2.5 (µg/m³)', 'PM2.5'),
        ('PM10 (µg/m³)', 'PM10'),
        ('Temperatura (°C)', 'Temperatura'),
        ('Humedad (%)', 'Humedad Relativa')
    ]
    
    for col, label in variables:
        if col in df.columns:
            values = df[col].dropna()
            if len(values) > 0:
                stats_data.append({
                    'Variable': label,
                    'Mínimo': round(values.min(), 2),
                    'Máximo': round(values.max(), 2),
                    'Promedio': round(values.mean(), 2),
                    'Mediana': round(values.median(), 2),
                    'Desv. Estándar': round(values.std(), 2),
                    'Registros': len(values)
                })
    
    return pd.DataFrame(stats_data)


def _style_worksheet(ws, title: str):
    """Aplica estilos a una hoja de Excel."""
    # Estilos
    header_fill = PatternFill(start_color="2C5AA0", end_color="2C5AA0", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14, color="1F4788")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Si hay filas, aplicar formato
    if ws.max_row > 0:
        # Título en la primera fila (si existe)
        if ws['A1'].value and isinstance(ws['A1'].value, str) and 'Reporte' in ws['A1'].value:
            ws['A1'].font = title_font
            try:
                ws.merge_cells(f'A1:{chr(64 + ws.max_column)}1')
            except:
                pass  # Si ya está fusionada, continuar
        
        # Encontrar fila de encabezado (buscar la primera fila con datos)
        header_row = 1
        for row in range(1, ws.max_row + 1):
            if ws.cell(row, 1).value is not None:
                header_row = row
                break
        
        # Aplicar estilo a encabezados
        for cell in ws[header_row]:
            # Saltar celdas fusionadas
            if hasattr(cell, 'column_letter'):
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
        
        # Aplicar bordes a todas las celdas con datos
        for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row, 
                                min_col=1, max_col=ws.max_column):
            for cell in row:
                # Saltar celdas fusionadas
                if hasattr(cell, 'column_letter'):
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-ajustar ancho de columnas
        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row_idx, col_idx)
                # Saltar celdas fusionadas
                if hasattr(cell, 'value') and cell.value:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width


def generate_excel_report(
    period: str,
    devices: Optional[List[str]] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    channels: Optional[List[SensorChannel]] = None,
    aggregate_by_minute: bool = True
) -> BytesIO:
    """
    Genera un reporte Excel para el período especificado.
    
    Args:
        period: Tipo de período ('hour', '24hours', 'month', 'year', 'custom')
        devices: Lista de device_ids a incluir (None = todos)
        start_date: Fecha de inicio (para custom)
        end_date: Fecha de fin (para custom)
        channels: Canales a incluir (None = ambos)
        aggregate_by_minute: Si True, agrega datos por minuto. Si False, datos crudos.
    
    Returns:
        BytesIO con el contenido del Excel
    """
    # Calcular rango de fechas según el período
    now = datetime.now(BOGOTA)
    today = now.date()
    
    if period == 'hour':
        # Última hora
        start_dt = now - timedelta(hours=1)
        end_dt = now
        title = "Reporte Última Hora - Minuto a Minuto"
    elif period == '24hours':
        # Últimas 24 horas
        start_dt = now - timedelta(hours=24)
        end_dt = now
        title = "Reporte Últimas 24 Horas - Minuto a Minuto"
    elif period == 'month':
        # Mes actual
        start_dt, _ = _bounds_of_day_local(date(today.year, today.month, 1))
        _, end_dt = _bounds_of_day_local(today)
        title = f"Reporte Mes {today.month}/{today.year} - Minuto a Minuto"
    elif period == '7days':
        # Últimos 7 días
        start_dt, _ = _bounds_of_day_local(today - timedelta(days=6))
        _, end_dt = _bounds_of_day_local(today)
        title = "Reporte Últimos 7 Días - Minuto a Minuto"
    elif period == 'year':
        # Año actual
        start_dt, _ = _bounds_of_day_local(date(today.year, 1, 1))
        _, end_dt = _bounds_of_day_local(today)
        title = f"Reporte Año {today.year} - Minuto a Minuto"
    elif period == 'custom' and start_date and end_date:
        start_dt, _ = _bounds_of_day_local(start_date)
        _, end_dt = _bounds_of_day_local(end_date)
        title = f"Reporte Personalizado - Minuto a Minuto"
    else:
        raise ValueError(f"Período no válido: {period}")
    
    # Obtener datos
    df = _get_measurements_df(start_dt, end_dt, devices, channels)
    
    if df.empty:
        # Crear Excel vacío con mensaje
        buffer = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte"
        ws['A1'] = "No se encontraron datos para el período seleccionado"
        ws['A1'].font = Font(bold=True, size=12)
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    # Agregar por minuto si se solicita
    if aggregate_by_minute:
        df_data = _aggregate_by_minute(df)
    else:
        df_data = df.copy()
    
    # Calcular estadísticas
    df_stats = _calculate_statistics_df(df_data)
    
    # Crear Excel
    buffer = BytesIO()
    wb = Workbook()
    
    # ========== HOJA 1: INFORMACIÓN GENERAL ==========
    ws_info = wb.active
    ws_info.title = "Información"
    
    info_data = [
        [title],
        [],
        ['Período:', f"{start_dt.strftime('%Y-%m-%d %H:%M')} - {end_dt.strftime('%Y-%m-%d %H:%M')}"],
        ['Generado:', datetime.now(BOGOTA).strftime('%Y-%m-%d %H:%M:%S')],
        ['Zona Horaria:', 'America/Bogota (UTC-5)'],
        ['Total de registros:', len(df_data)],
        [],
        ['Dispositivos incluidos:'],
    ]
    
    if devices:
        for dev in devices:
            info_data.append(['', f'• {label_for(dev)} ({dev})'])
    else:
        info_data.append(['', '• Todos los dispositivos'])
    
    info_data.append([])
    info_data.append(['Canales incluidos:'])
    if channels:
        for ch in channels:
            info_data.append(['', f'• {ch.name}'])
    else:
        info_data.append(['', '• Um1 y Um2'])
    
    for row_data in info_data:
        ws_info.append(row_data)
    
    _style_worksheet(ws_info, title)
    
    # ========== HOJA 2: ESTADÍSTICAS ==========
    if not df_stats.empty:
        ws_stats = wb.create_sheet("Estadísticas")
        ws_stats.append(['Estadísticas Generales'])
        ws_stats.append([])
        
        for r in dataframe_to_rows(df_stats, index=False, header=True):
            ws_stats.append(r)
        
        _style_worksheet(ws_stats, "Estadísticas")
    
    # ========== HOJA 3: DATOS COMPLETOS ==========
    ws_data = wb.create_sheet("Datos Completos")
    ws_data.append([f'{title} - Todos los Registros'])
    ws_data.append([])
    
    for r in dataframe_to_rows(df_data, index=False, header=True):
        ws_data.append(r)
    
    _style_worksheet(ws_data, title)
    
    # ========== HOJA 4: RESUMEN POR DISPOSITIVO ==========
    ws_summary = wb.create_sheet("Resumen Dispositivos")
    ws_summary.append(['Resumen por Dispositivo'])
    ws_summary.append([])
    
    summary_data = df_data.groupby('Dispositivo').agg({
        'PM2.5 (µg/m³)': ['mean', 'min', 'max', 'count'],
        'PM10 (µg/m³)': ['mean', 'min', 'max'],
        'Temperatura (°C)': ['mean', 'min', 'max'],
        'Humedad (%)': ['mean', 'min', 'max'],
    }).round(2)
    
    # Aplanar columnas multi-nivel
    summary_data.columns = ['_'.join(col).strip() for col in summary_data.columns.values]
    summary_data = summary_data.reset_index()
    
    # Renombrar columnas
    summary_data.columns = [
        'Dispositivo',
        'PM2.5 Promedio', 'PM2.5 Mínimo', 'PM2.5 Máximo', 'Total Registros',
        'PM10 Promedio', 'PM10 Mínimo', 'PM10 Máximo',
        'Temp Promedio', 'Temp Mínimo', 'Temp Máximo',
        'HR Promedio', 'HR Mínimo', 'HR Máximo'
    ]
    
    for r in dataframe_to_rows(summary_data, index=False, header=True):
        ws_summary.append(r)
    
    _style_worksheet(ws_summary, "Resumen")
    
    # Guardar
    wb.save(buffer)
    buffer.seek(0)
    
    logger.info(f"Reporte Excel generado: {period}, {len(df_data)} registros")
    return buffer
